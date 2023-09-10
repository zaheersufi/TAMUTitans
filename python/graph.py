import numpy as np
import matplotlib.pyplot as plt
import openpyxl


def read_column_a(file_name):
    # Load the workbook
    wb = openpyxl.load_workbook(file_name)

    ws = wb.active

    column_a_values = []

    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
        cell_value = row[0]  # Since we are reading a single column
        if cell_value is not None:
            column_a_values.append(cell_value)

    return column_a_values


def read_column_b(file_name):
    wb = openpyxl.load_workbook(file_name)

    ws = wb.active

    column_b_values = []

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row, values_only=True):
        cell_value = row[0]  # Since we are reading a single column
        if cell_value is not None:
            column_b_values.append(cell_value)

    return column_b_values


def read_column_c(file_name):
    wb = openpyxl.load_workbook(file_name)

    ws = wb.active

    column_c_values = []

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=ws.max_row, values_only=True):
        cell_value = row[0]  # Since we are reading a single column
        if cell_value is not None:
            column_c_values.append(cell_value)

    return column_c_values


file_name = 'mydata.xlsx'
values_in_column_a = read_column_a(file_name)
values_in_column_b = read_column_b(file_name)
values_in_column_c = read_column_c(file_name)

# {"9/12/2023": [0, 0, 0, 2]} #Personal, School, Food, Misc
def fill_by_dates(a, b, c):
    arrayIndex = -1
    if c == "Personal":
        arrayIndex = 0
    elif c == "School":
        arrayIndex = 1
    elif c == "Food":
        arrayIndex = 2
    elif c == "Miscellaneous":
        arrayIndex = 3
    dict = {}
    for day in a:
        if day not in dict:
            dict[day] = (0, 0, 0, 0)
            dict[day][arrayIndex] = b
        else:
            dict[day][arrayIndex] = b
    return dict

all_values = fill_by_dates(values_in_column_a, values_in_column_b, values_in_column_c)
print(all_values)

dates = list(values_in_column_a)
values = list(values_in_column_b)

fig = plt.figure(figsize=(10, 5))

# creating the bar plot
plt.bar(dates, values, color='maroon',
        width=0.4)

plt.xlabel("Date")
plt.ylabel("Amounts")
plt.title("Budget Chart")
plt.show()
