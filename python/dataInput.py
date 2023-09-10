import openpyxl


def add_expense(date, category, amount):
    wb = openpyxl.load_workbook('mydata.xlsx')
    ws = wb.active

    last_row = ws.max_row

    new_row = last_row + 1

    ws[f'A{new_row}'] = date
    ws[f'B{new_row}'] = category
    ws[f'C{new_row}'] = amount

    wb.save('mydata.xlsx')


date = str(input("Enter the date: "))
cat = input("Enter the category: ")
amount = int(input("Enter the amount: "))

add_expense(date, cat, amount)
